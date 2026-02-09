import re
import tempfile
import pdfplumber
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def extract_text_from_pdf_bytesio(pdf_bytesio):
    pdf_bytesio.seek(0)
    with pdfplumber.open(pdf_bytesio) as pdf:
        if len(pdf.pages) > 0:
            first_page = pdf.pages[0]
            return first_page.extract_text() or ""
    return ""

def normalize(text):
    if not text:
        return ""
    try:
        text = str(text)
        replacements = [
            ('\xa0', ' '),
            ('\u00a0', ' '),
            ('\u2013', '-'),
            ('\u2014', '-'),
            ('\u201c', '"'),
            ('\u201d', '"'),
            ('\u2018', "'"),
            ('\u2019', "'"),
            ('\u2026', '...'),
            ('\u00b0', '°'),
            ('\u00b2', '²'),
            ('\u00b3', '³'),
            ('：', ':'),
            ('–', '-'),
            ('—', '-'),
        ]
        
        for old, new in replacements:
            text = text.replace(old, new)        
        text = re.sub(r'\n+', ' ', text)
        text = re.sub(r'\s+', ' ', text)        
        return text.strip()
    except Exception as e:
        print(f"Warning: Error in normalize function: {e}")
        return str(text) if text else ""

def smart_split_applicant_block(full_block):
    block = full_block.strip().replace("  ", " ")
    address_keywords = [
        "Door No", "Old No", "New No", "Plot No", "Flat No",
        r"No\.", r"No\\s", "Street", "Salai", "Nagar", "Colony", 
        "Village", "Complex", "Avenue", r"[0-9]{6}"]
    lines = [line.strip(" ,:-") for line in block.splitlines() if line.strip()]
    num_lines = len(lines)
    if num_lines == 2:
        return lines[0], lines[1]
    if num_lines >= 3:
        name_lines, address_lines = [], []
        for line in lines:
            if any(re.search(kw, line, re.IGNORECASE) for kw in address_keywords) or re.match(r"^\d", line):
                address_lines.append(line)
            else:
                name_lines.append(line)
        if not address_lines and len(name_lines) > 1:
            address_lines.append(name_lines.pop())
        name = " ".join(name_lines).strip(" ,:-") if name_lines else "Not Found"
        address = " ".join(address_lines).strip(" ,:-") if address_lines else "Not Found"
        return name, address
    if num_lines == 1:
        single_line = lines[0]
        address_start_pattern = re.search(
            r"\b(\d{1,3}(?:/\d{1,3})?|Door\s+No\.?|Old\s+No\.?|New\s+No\.?|Plot\s+No\.?|Flat\s+No\.?|No\.?|Street|Salai|Nagar|Colony|Village|Complex|Avenue|[0-9]{6})\b",
            single_line, re.IGNORECASE)
        if address_start_pattern:
            idx = address_start_pattern.start()
            name = single_line[:idx].strip(" ,:-")
            address = single_line[idx:].strip(" ,:-")
            return name, address
        if "," in single_line:
            parts = [p.strip(" ,:-") for p in single_line.rsplit(",", 1)]
            if len(parts) == 2:
                return parts[0], parts[1]
        words = re.findall(r"\S+", single_line)
        if len(words) > 1:
            last_word_clean = words[-1].rstrip(",.:;-")
            name = " ".join(words[:-1]).strip(" ,:-")
            address = last_word_clean.strip(" ,:-")
            return name, address
        return single_line.strip(" ,:-"), "Not Found"
    return block.strip(" ,:-"), "Not Found"

def extract_area_name(site_address):
    if not site_address or site_address == "Not Found":
        return "Not Found"
    def clean_prefix(text):
        return re.sub(
            r"^(?:ward\s*[-]?\s*\w*\s*of\s+|[A-Za-z]\s+of\s+|of\s+|situated\s+in\s+|in\s+)",
            "",
            text,
            flags=re.IGNORECASE
        ).strip()
    village_match = re.search(r"\b([A-Za-z.\s-]+?)\s+Village\b", site_address, re.IGNORECASE)
    if village_match:
        return clean_prefix(village_match.group(1).strip())
    taluk_match = re.search(r"\b([A-Za-z.\s-]+?)\s+Taluk\b", site_address, re.IGNORECASE)
    if taluk_match:
        return clean_prefix(taluk_match.group(1).strip())
    chennai_match = re.search(r",\s*([A-Za-z.\s-]+?),\s*Chennai", site_address, re.IGNORECASE)
    if chennai_match:
        return clean_prefix(chennai_match.group(1).strip())
    pin_match = re.search(r",\s*([A-Za-z.\s-]+?)\s*-?\s*\d{6}", site_address)
    if pin_match:
        return clean_prefix(pin_match.group(1).strip())
    return clean_prefix(site_address.strip())

def extract_fields(text):
    try:
        text = normalize(text)
        fields = {}
        def search_field(label, pattern):
            match = re.search(pattern, text, re.IGNORECASE)
            if not match:
                return "Not Found"
            return next((g.strip() for g in match.groups() if g), "Not Found")
        fields["File No."] = search_field("File No.", r"File\s*No\.?\s*[:\-]?\s*(CMDA[^\s]+)")
        fields["Planning Permission No."] = search_field("Planning Permission No.", r"Planning\s*Permission\s*No\.?\s*[:\-]?\s*([A-Z0-9/\-]+)")
        fields["Permit No."] = search_field("Permit No.", r"Permit\s*No\.?\s*[:\-]?\s*([A-Z0-9/\-]+)")
        fields["Date of permit"] = search_field(
            "Date of permit",
            r"(?:([0-9]{2}[-/][0-9]{2}[-/][0-9]{4})\s*Date\s*of\s*permit|Date\s*of\s*permit\s*[:\-]?\s*([0-9]{2}[-/][0-9]{2}[-/][0-9]{4}))"
        )
        fields["Date of Application"] = search_field(
            "Date of Application",
            r"(?:Date\s*of\s*Application\s*[:\-]?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{4})|(\d{1,2}[-/]\d{1,2}[-/]\d{4})\s*Date\s*of\s*Application)"
        )
        fields["Mobile No."] = search_field("Mobile No.", r"Mobile\s*No\.?\s*[:\-]?\s*(\d{10})")
        fields["Email ID"] = search_field("Email ID", r"Email\s*ID\s*[:\-]?\s*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})")
        applicant_block = re.search(
            r"Name\s+of\s+Applicant\s+with\s+Address\s*[:\-]?\s*(.+?)\s+(Mobile\s*No|Email\s*ID|Date\s+of\s+Application|Date\s+of\s*permit)",
            text, re.IGNORECASE | re.DOTALL
        )
        if applicant_block:
            full_block = applicant_block.group(1).strip()
            name, address = smart_split_applicant_block(full_block)
            fields["Applicant Name"] = name
            fields["Applicant Address"] = address
        else:
            fields["Applicant Name"] = "Not Found"
            fields["Applicant Address"] = "Not Found"
        nature_block = re.search(
            r"Nature\s+of\s+Development\s*[:\-]?\s*(.*?)(?=\s+Site\s+Address\s*[:\-]?)",
            text, re.IGNORECASE | re.DOTALL
        )
        site_block = re.search(
            r"Site\s+Address\s*[:\-]?\s*(.*?)(?=Development\s+Charge|Receipt|Permission\s+is\s+granted|Signature|Yours\s+faithfully|The\s+permit\s+expires|\Z)",
            text, re.IGNORECASE | re.DOTALL
        )
        nature_text = normalize(nature_block.group(1)) if nature_block else "Not Found"
        fields["Nature of Development"] = nature_text
        dwelling_match = re.search(
            r"\b\d*\s*dwellings?\s+units?\b|"           
            r"\b\d*\s*dwelling\s+unit\b|"              
            r"\bsingle\s+dwelling\s+unit\b|"           
            r"\b\d*\s*dwellings?\b|"                   
            r"\b\d*\s*dwelling\b",                     
            nature_text,
            re.IGNORECASE
        )
        fields["Dwelling Unit Info"] = dwelling_match.group(0) if dwelling_match else ""
        site_address = normalize(site_block.group(1)) if site_block else "Not Found"
        fields["Site Address"] = site_address
        fields["Area Name"] = extract_area_name(site_address)
        return fields
        
    except Exception as e:
        print(f"Error in extract_fields: {e}")
        default_fields = {
            "File No.": "Error",
            "Planning Permission No.": "Error",
            "Permit No.": "Error",
            "Date of permit": "Error",
            "Date of Application": "Error",
            "Mobile No.": "Error",
            "Email ID": "Error",
            "Applicant Name": "Error",
            "Applicant Address": "Error",
            "Nature of Development": "Error",
            "Dwelling Unit Info": "",
            "Site Address": "Error",
            "Area Name": "Error"
        }
        return default_fields

def export_to_xlsx(data_list, year, urls, approved_links, approved_letter, architect_details):
    try:
        downloads_folder = Path.home() / "Downloads"
        downloads_folder.mkdir(parents=True, exist_ok=True)
        download_path = downloads_folder / f"CMDA_{year}.xlsx"
        temp_dir = Path(tempfile.gettempdir())
        temp_path = temp_dir / f"CMDA_{year}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Permit Data"
        FIELD_ORDER = [
            "File No.", "Planning Permission No.", "Permit No.",
            "Date of permit", "Date of Application",
            "Mobile No.", "Email ID",
            "Applicant Name", "Applicant Address",
            "Nature of Development", "Dwelling Unit Info",
            "Site Address", "Area Name"
        ]
        headers = FIELD_ORDER + [
            "Architect Name", "Architect Address", "Architect Email", "Architect Mobile",
            "View Online", "Approved Plan", "Approval Letter"
        ]
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        combined_data = list(zip(data_list, urls, approved_links, approved_letter, architect_details))
        combined_data.sort(key=lambda x: x[0].get("Area Name", "").strip().lower())
        for row_idx, (data, url, approved_url, letter_url, architect) in enumerate(combined_data, start=2):
            for col_idx, key in enumerate(FIELD_ORDER, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(data.get(key, "")))
            ws.cell(row=row_idx, column=len(FIELD_ORDER) + 1, value=architect.get("name", ""))
            ws.cell(row=row_idx, column=len(FIELD_ORDER) + 2, value=architect.get("address", ""))
            ws.cell(row=row_idx, column=len(FIELD_ORDER) + 3, value=architect.get("email", ""))
            ws.cell(row=row_idx, column=len(FIELD_ORDER) + 4, value=architect.get("mobile", ""))
            link_cell = ws.cell(row=row_idx, column=len(FIELD_ORDER) + 5, value="View PDF")
            if url:
                link_cell.hyperlink = url
                link_cell.style = "Hyperlink"
            approved_cell = ws.cell(row=row_idx, column=len(FIELD_ORDER) + 6, value="View Approved Plan")
            if approved_url:
                approved_cell.hyperlink = approved_url
                approved_cell.style = "Hyperlink"
            letter_cell = ws.cell(row=row_idx, column=len(FIELD_ORDER) + 7, value="View Approval Letter")
            if letter_url:
                letter_cell.hyperlink = letter_url
                letter_cell.style = "Hyperlink"
        wb.save(download_path)
        wb.save(temp_path)
        crm_url = f"file:///{temp_path.as_posix()}"
        return str(temp_path), str(download_path)
    except Exception as e:
        print(f"Error in export_to_xlsx: {e}")
        downloads_folder = Path.home() / "Downloads"
        download_path = downloads_folder / f"CMDA_{year}_ERROR.xlsx"
        temp_dir = Path(tempfile.gettempdir())
        temp_path = temp_dir / f"CMDA_{year}_ERROR.xlsx"        
        wb = Workbook()
        ws = wb.active
        ws.title = "Error Report"
        ws['A1'] = "Error during export"
        ws['A2'] = str(e)
        wb.save(download_path)
        wb.save(temp_path)
        return str(temp_path), str(download_path)